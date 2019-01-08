from secondary_func import*
from datetime import timedelta
from datetime import datetime
import sys
import openpyxl
wb = openpyxl.Workbook()
ws = wb.active
full_info = {}
sign = 0
time_str = '00:00:00'
delta_work = datetime.strptime(time_str,"%H:%M:%S")
work_time = timedelta(hours=8)
time_comparison = timedelta(seconds=0)
recycling = datetime.strptime(time_str,"%H:%M:%S")
f = open('tst','r')
for line in f:
    Y = line[8:29]
    date_and_time = datetime.strptime(line[10:29], "%Y-%m-%d %H:%M:%S")
    date = line[10:20]
    U = int(Y[0])
    family = replace_name(U)
    if not  date in full_info.keys():
        full_info[date] = {}
    if U in full_info[date].keys():
        if full_info[date][U][0] < date_and_time:
            full_info[date][U][0] = date_and_time
        if full_info[date][U][1] > date_and_time:
            full_info[date][U][1] = date_and_time
    else:
        full_info[date][U] = [date_and_time, date_and_time,delta_work,recycling,sign]
number_family = [1,2,3,4,5,6,7,8]
year = input("Get a year: ")
month = input("Get a month: ")
last_day = monthrange(int(year), int(month))[1]
a = datetime(int(year), int(month), 1)
b = datetime(int(year), int(month), last_day)
for dt in daterange(a, b, inclusive=True):
  date = dt.strftime("%Y-%m-%d")
  if date in full_info.keys():
    for i in full_info[date].keys():
      name = replace_name(i)
      if name != 'Рощин':
        if full_info[date][i][0] == full_info[date][i][1]:
          print("AHTUNG! " + name + " have only one fingerprint", file=sys.stderr)
          print('Фамилия', name)
          print('Дата время в системе',full_info[date][i][0])
          print('\n\n\tВыберете вариант:\n\n\t1.Ввести время прихода\n\t2.Введите время ухода')
          l = 0
          while l!=1:
            time_inp = input('\n\n\tВведите пункт меню:')
            if time_inp == '1' or time_inp == '2':
              l = 1
              if time_inp == '1':
                time_begin = input('Введите время прихода в формате час*ПРОБЕЛ*минуты*секунды(если есть)')
                time_begin_str = date + ' ' + time_begin
                time_begin_z = datetime.strptime(time_begin_str,"%Y-%m-%d %H %M %S")
                full_info[date][i][0] = time_begin_z
              elif time_inp == '2': 
                time_end = input('Введите время ухода в формате час*ПРОБЕЛ*минуты*секунды(если есть)')
                time_end_str = date + ' ' + time_end
                time_end_z = datetime.strptime(time_end_str,"%Y-%m-%d %H %M %S")
                full_info[date][i][1] = time_end_z
              else:
                pass
              print('после измений')
              work_delta = full_info[date][i][1] - full_info[date][i][0]
              full_info[date][i][2] = work_delta
              recycling = work_delta - work_time
              full_info[date][i][3] = recycling
              print(name,' ',date,' ',str(work_delta),' переработка: ',recycling)
          else:
            print('Введите правильный пункт меню!')
        else:
          work_delta = full_info[date][i][0] - full_info[date][i][1]
          full_info[date][i][2] = work_delta
          recycling = work_delta - work_time        
          if recycling < time_comparison:
            full_info[date][i][4]= 1
            full_info[date][i][3] = abs(recycling)
          else:
            full_info[date][i][3] = recycling
      else:
        pass

print('\n\n\n\n\tТестовое количество значений\n\n\n\n\n',len(full_info))      
for dt in daterange(a, b, inclusive=True):
    date = dt.strftime("%Y-%m-%d")
    if date in full_info.keys():
        for i in full_info[date].keys():
            name = replace_name(i)
            for row in range(2,84):
                rec_name = ws.cell(column = 1, row = row, value = name)
                rec_date = ws.cell(column = 2, row = row, value = date)
                rec_time_start = ws.cell(column = 3, row = row, value = full_info[date][i][0])
                rec_time_end = ws.cell(column = 4, row = row, value = full_info[date][i][1])
                rec_delta = ws.cell(column = 5, row = row, value = full_info[date][i][2])
                rec_recycling = ws.cell(column = 6, row = row, value = full_info[date][i][3])
            

wb.save("табель_тест_1.xlsx")
#-------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
list_print_1 = []
list_print_2 = []
for dt in daterange(a, b, inclusive=True):
    date = dt.strftime("%Y-%m-%d")
    if date in full_info.keys():
        for i in full_info[date].keys():
            name = replace_name(i)
            if name != 'Рощин':
                if full_info[date][i][4] == 0:
                    prt_1 = str(name)+ ' ' + str(date) + ' '+ str(full_info[date][i][2])+ ' ' + ' переработка: ' + str(full_info[date][i][3])
                else:
                    prt_1 = str(name)+ ' ' + str(date) + ' '+ str(full_info[date][i][2])+ ' ' + ' недоработка: ' + str(abs(full_info[date][i][3]))
                list_print_1.append(prt_1)
            else:
                prt_2 = str(name)+ ' ' + str(date) + ' '+ 'время прихода' + str(full_info[date][i][0])
                list_print_2.append(prt_2)
for i in list_print_1:
    if not 'Еремин' in i:
        if 'переработка:' in i:
            print(i)
        else:
            print(i,file=sys.stderr)
    else:
        pass
print('---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------')
for n in list_print_2:
    print(n)
print('FINAL!')
