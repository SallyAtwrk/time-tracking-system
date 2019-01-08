from secondary_func import*
from datetime import timedelta
from datetime import datetime
import sys
import openpyxl
wb = openpyxl.Workbook()
ws = wb.active
ws.cell(column = 1, row = 1, value = 'Фамилия')
ws.cell(column = 2, row = 1, value = 'Дата')
ws.cell(column = 3, row = 1, value = 'Отметка входа')
ws.cell(column = 4, row = 1, value = 'Отметка выхода')
ws.cell(column = 5, row = 1, value = 'Общее время работы')
ws.cell(column = 6, row = 1, value = 'Переботка')
ws.cell(column = 7, row = 1, value = 'Примечания')
full_info = {}
sign = 0 #Хрен знает, что это за переменная и зачем я пишу её в объект. Возможно это знак, определяющий дельту - если положительный значит переработка, если отрицательный - то недоработка
time_str = '00:00:00'
count = 2 #Счетчик. Через него строится запись в колонки и столбцы. 
delta_work = datetime.strptime(time_str,"%H:%M:%S")# Реально отработанное время
work_time = timedelta(hours=9)# Длительность рабочего дня
time_comparison = timedelta(seconds=0) #Разница между реально отработанным временем и длительностью рабочего дня
recycling = datetime.strptime(time_str,"%H:%M:%S")# Переработка
f = open('4_attlog','r')# Иеальные данные tst_0/Не идеальные данные tst_1 Загрузка файла с данными
for line in f:# Перебор всех строк в файле с данными
    Y = line[8:29]
    date_and_time = datetime.strptime(line[10:29], "%Y-%m-%d %H:%M:%S")
    date = line[10:20]
    U = int(Y[0])
    family = replace_name(U)
    if not  date in full_info.keys():
        full_info[date] = {}
    if U in full_info[date].keys():
        if full_info[date][U][0] < date_and_time:
            full_info[date][U][0] = date_and_time
        if full_info[date][U][1] > date_and_time:
            full_info[date][U][1] = date_and_time
    else:
        full_info[date][U] = [date_and_time, date_and_time,delta_work,recycling,sign]
number_family = [1,2,3,4,5,6,7,8] ; #id работников 
year = input("Get a year: ") # ввод года для построения табеля
month = input("Get a month: ") # ввод месяц для построения табеля
last_day = monthrange(int(year), int(month))[1] # Определение последнего дня месяца? 
a = datetime(int(year), int(month), 1) ;
b = datetime(int(year), int(month), last_day)
for dt in daterange(a, b, inclusive=True):
  date = dt.strftime("%Y-%m-%d")
  if date in full_info.keys():
    for i in full_info[date].keys():
      name = replace_name(i)
      rec_name = ws.cell(column = 1, row = count, value = name)
      rec_date = ws.cell(column = 2, row = count, value = date)
      if name != 'Рощин':
        if full_info[date][i][0] == full_info[date][i][1]:
          print("AHTUNG! " + name + " have only one fingerprint", file=sys.stderr)
          print('Фамилия', name)
          print('Дата время в системе',full_info[date][i][0])
          print('\n\n\tВыберете вариант:\n\n\t1.Ввести время прихода\n\t2.Введите время ухода')
          l = 0
          while l!=1:
            time_inp = input('\n\n\tВведите пункт меню:')
            if time_inp == '1' or time_inp == '2':
              l = 1
              if time_inp == '1':
                time_begin = input('Введите время прихода в формате час*ПРОБЕЛ*минуты*секунды(если есть)')
                time_begin_str = date + ' ' + time_begin
                time_begin_z = datetime.strptime(time_begin_str,"%Y-%m-%d %H %M %S")
                full_info[date][i][0] = time_begin_z
              elif time_inp == '2': 
                time_end = input('Введите время ухода в формате час*ПРОБЕЛ*минуты*секунды(если есть)')
                time_end_str = date + ' ' + time_end
                time_end_z = datetime.strptime(time_end_str,"%Y-%m-%d %H %M %S")
                full_info[date][i][1] = time_end_z
              else:
                pass
              print('после измений')
              value_for_record = datetime.time(full_info[date][i][0])
              rec_time_start = ws.cell(column = 3, row = count, value = value_for_record)
              value_for_record = datetime.time(full_info[date][i][1])
              rec_time_end = ws.cell(column = 4, row = count, value = value_for_record)
              work_delta = (full_info[date][i][1] - full_info[date][i][0]).seconds
              print(work_delta, 'Проверка 1. Рабочее время, секунды')
              full_info[date][i][2] = work_delta
              rec_delta = ws.cell(column = 5, row = count, value = full_info[date][i][2])
              print(type(work_delta),'work_delta',type(work_time),'work_time')
              recycling = work_delta - work_time.seconds
              recycling_hour = recycling/3600
              print(recycling_hour,'- целочисленное значение часов переработки')
              print(recycling, 'Проверка 2. Переработка, секнуды')
              full_info[date][i][3] = recycling
              rec_recycling = ws.cell(column = 6, row = count, value = full_info[date][i][3])
              print(name,' ',date,' ',str(work_delta),' переработка: ',recycling)
          else:
            print('Введите правильный пункт меню!')
        else:
          work_delta = full_info[date][i][0] - full_info[date][i][1]#Короче весь этот блок надо переделывать. Полность. Считать через секунды и формировать значения из  секунд. Еще надо сменить название переменной переботки на более корректное. Добавить id новых сотрудинов, выкинуть Федоровича из программы. Разобраться с форматированнием вывода времени и даты. Разобраться с ранжированием месяца.
          value_for_record = datetime.time(full_info[date][i][0])
          rec_time_start = ws.cell(column = 4, row = count, value = value_for_record)
          value_for_record = datetime.time(full_info[date][i][1])
          rec_time_end = ws.cell(column = 3, row = count, value = value_for_record)
          full_info[date][i][2] = work_delta
          rec_delta = ws.cell(column = 5, row = count, value = full_info[date][i][2])
          recycling = work_delta - work_time
          if recycling < time_comparison:
              full_info[date][i][4]= 1
              full_info[date][i][3] = abs(recycling)
              rec_recycling = ws.cell(column = 6, row = count, value = full_info[date][i][3])
          else:
              full_info[date][i][3] = recycling
              rec_recycling = ws.cell(column = 6, row = count, value = full_info[date][i][3])
          count = count +1
      else:
          value_for_record = datetime.time(full_info[date][i][0])
          rec_time_start = ws.cell(column = 3, row = count, value = value_for_record)
          count = count +1
    else:
        pass
    

wb.save("табель_апрель.xlsx")
#-------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
list_print_1 = []
list_print_2 = []
for dt in daterange(a, b, inclusive=True):
    date = dt.strftime("%Y-%m-%d")
    if date in full_info.keys():
        for i in full_info[date].keys():
            name = replace_name(i)
            if name != 'Рощин':
                if full_info[date][i][4] == 0:
                    prt_1 = str(name)+ ' ' + str(date) + ' '+ str(full_info[date][i][2])+ ' ' + ' переработка: ' + str(full_info[date][i][3])
                else:
                    prt_1 = str(name)+ ' ' + str(date) + ' '+ str(full_info[date][i][2])+ ' ' + ' недоработка: ' + str(abs(full_info[date][i][3]))
                list_print_1.append(prt_1)
            else:
                prt_2 = str(name)+ ' ' + str(date) + ' '+ 'время прихода' + str(full_info[date][i][0])
                list_print_2.append(prt_2)
for i in list_print_1:
    if not 'Еремин' in i:
        if 'переработка:' in i:
            print(i)
        else:
            print(i,file=sys.stderr)
    else:
        pass
print('---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------')
for n in list_print_2:
    print(n)
print('FINAL!')
