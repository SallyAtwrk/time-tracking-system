from secondary_func import*
from datetime import timedelta
from datetime import datetime
import time
import sys
import openpyxl
wb = openpyxl.Workbook()
ws = wb.active
ws.cell(column = 1, row = 1, value = 'Фамилия')
ws.cell(column = 2, row = 1, value = 'Дата')
ws.cell(column = 3, row = 1, value = 'Отметка входа')
ws.cell(column = 4, row = 1, value = 'Отметка выхода')
ws.cell(column = 5, row = 1, value = 'Общее время работы')
ws.cell(column = 6, row = 1, value = 'Переботка')
ws.cell(column = 7, row = 1, value = 'Примечания')
full_info = {}
time_str = '00:00:00'
count = 2 #Начальное значение счетчика. С помощью этой переменной определяется значеение счетчика 
delta_work = datetime.strptime(time_str,"%H:%M:%S")# Реально отработанное время
work_time = timedelta(hours=8)# Длительность рабочего дня
time_comparison = timedelta(seconds=0) #Разница между реально отработанным временем и длительностью рабочего дня
recycling_seconds = 0
recycling_hour = 0.0
recycling_hour_round = 0.0
recycling_data = time.strftime('%H:%M:%S', time.gmtime(abs(recycling_seconds)))
money_rate_8h = 1000.0
wage = 0.0
f = open('tst_0','r')# Иеальные данные tst_0/Не идеальные данные tst_1 Загрузка файла с данными
for line in f:# Перебор всех строк в файле с данными
    Y = line[8:29]
    date_and_time = datetime.strptime(line[10:29], "%Y-%m-%d %H:%M:%S")
    date = line[10:20]
    U = int(Y[0])
    family = replace_name(U)
    if not  date in full_info.keys():
        full_info[date] = {}
    if U in full_info[date].keys():
        if full_info[date][U][0] < date_and_time:
            full_info[date][U][0] = date_and_time
        if full_info[date][U][1] > date_and_time:
            full_info[date][U][1] = date_and_time
    else:
        full_info[date][U] = [date_and_time, date_and_time,delta_work,recycling_seconds,recycling_hour,recycling_hour_round,recycling_data,wage]
number_family = [1,2,3,4,5,6,7,8,9,10] ; #id работников 
year = input("Get a year: ") # ввод года для построения табеля
month = input("Get a month: ") # ввод месяц для построения табеля
last_day = monthrange(int(year), int(month))[1]
a = datetime(int(year), int(month), 1) ;
b = datetime(int(year), int(month), last_day)
for dt in daterange(a, b, inclusive=True):
  date = dt.strftime("%Y-%m-%d")
  if date in full_info.keys():
    for i in full_info[date].keys():
      name = replace_name(i)
      rec_name = ws.cell(column = 1, row = count, value = name)
      rec_date = ws.cell(column = 2, row = count, value = date)
      if name != 'Рощин':
        if full_info[date][i][0] == full_info[date][i][1]:
          print("ПРЕДУПРЕЖДЕНИЕ! " + name + " имеет только одну отметку в рабочем дне!", file=sys.stderr)
          print('Дата время в системе',full_info[date][i][0])
          print('\n\n\tВыберете вариант:\n\n\t1.Ввести время прихода\n\t2.Введите время ухода')
          l = 0
          while l!=1:
            time_inp = input('\n\n\tВведите пункт меню:')
            if time_inp == '1' or time_inp == '2':
              l = 1
              if time_inp == '1':
                time_begin = input('Введите время прихода в формате час*ПРОБЕЛ*минуты*секунды(если есть)')
                time_begin_str = date + ' ' + time_begin
                time_begin_z = datetime.strptime(time_begin_str,"%Y-%m-%d %H %M %S")
                full_info[date][i][0] = time_begin_z
              elif time_inp == '2': 
                time_end = input('Введите время ухода в формате час*ПРОБЕЛ*минуты*секунды(если есть)')
                time_end_str = date + ' ' + time_end
                time_end_z = datetime.strptime(time_end_str,"%Y-%m-%d %H %M %S")
                full_info[date][i][1] = time_end_z
              else:
                pass
              print('после измений')
              value_for_record = datetime.time(full_info[date][i][0])
              rec_time_start = ws.cell(column = 3, row = count, value = value_for_record)
              value_for_record = datetime.time(full_info[date][i][1])
              rec_time_end = ws.cell(column = 4, row = count, value = value_for_record)
              work_delta = full_info[date][i][1] - full_info[date][i][0]
              work_delta_seconds = (full_info[date][i][1] - full_info[date][i][0]).seconds
              full_info[date][i][2] = work_delta
              rec_delta = ws.cell(column = 5, row = count, value = full_info[date][i][2])
              recycling_seconds = work_delta_seconds - work_time.seconds
              recycling_hour = recycling_seconds/3600
              recycling_hour_round = round(recycling_hour,2)
              if recycling_seconds >= 0:
                full_info[date][i][3] = recycling_seconds
                full_info[date][i][4] = recycling_hour
                full_info[date][i][5] = recycling_hour_round
                rec_recycling = ws.cell(column = 6, row = count, value = full_info[date][i][5])
              else:
                full_info[date][i][3] = recycling_seconds
                full_info[date][i][4] = recycling_hour
                full_info[date][i][5] = recycling_hour_round
                rec_recycling = ws.cell(column = 6, row = count, value = full_info[date][i][5])
          else:
            print('Введите правильный пункт меню!')
        else:
          work_delta = full_info[date][i][0] - full_info[date][i][1]
          work_delta_seconds = work_delta.seconds
          value_for_record = datetime.time(full_info[date][i][0])
          rec_time_start = ws.cell(column = 4, row = count, value = value_for_record)
          value_for_record = datetime.time(full_info[date][i][1])
          rec_time_end = ws.cell(column = 3, row = count, value = value_for_record)
          full_info[date][i][2] = work_delta
          rec_delta = ws.cell(column = 5, row = count, value = full_info[date][i][2])
          recycling_seconds = work_delta_seconds - work_time.seconds
          money_rate_1s = money_rate_8h/(8*3600)
          money_rate_1h = money_rate_8h/8
          wage_recycling_seconds = recycling_seconds*money_rate_1s
          recycling_hour = recycling_seconds/3600
          recycling_hour_round = round(recycling_hour,2)
          wage_recycling_hour = recycling_hour*money_rate_1h
          wage_recycling_hour_round = recycling_hour_round*money_rate_1h
          recycling_data = time.strftime('%H:%M:%S', time.gmtime(abs(recycling_seconds)))
          full_info[date][i][3] = recycling_seconds
          full_info[date][i][4] = recycling_hour
          full_info[date][i][5] = recycling_hour_round
          full_info[date][i][6] = recycling_data
          full_info[date][i][7] = wage_recycling_seconds
          rec_recycling = ws.cell(column = 6, row = count, value = full_info[date][i][6])
          #print('Зарплата через секунды:',wage_recycling_seconds,'\n\n\nЗарплата через часы без округления:',wage_recycling_hour,'\n\n\nЗарплата через часы с округлением:',wage_recycling_hour_round)
          count = count +1
      else:
          value_for_record = datetime.time(full_info[date][i][0])
          rec_time_start = ws.cell(column = 3, row = count, value = value_for_record)
          count = count +1
    else:
        pass
    

wb.save("Проверочная таблица.xlsx")
#-------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
list_print_1 = []
list_print_2 = []
for dt in daterange(a, b, inclusive=True):
    date = dt.strftime("%Y-%m-%d")
    if date in full_info.keys():
        for i in full_info[date].keys():
            name = replace_name(i)
            if name != 'Рощин':
                if full_info[date][i][3] >= 0:
                    prt_1 = str(name)+ ' ' + str(date) + ' '+ str(full_info[date][i][2])+ ' ' + ' переработка: ' + str(full_info[date][i][6])
                else:
                    prt_1 = str(name)+ ' ' + str(date) + ' '+ str(full_info[date][i][2])+ ' ' + ' недоработка: ' + str(full_info[date][i][6])
                list_print_1.append(prt_1)
            else:
                prt_2 = str(name)+ ' ' + str(date) + ' '+ 'время прихода' + str(full_info[date][i][0])
                list_print_2.append(prt_2)
for i in list_print_1:
    if not 'Еремин' in i:
        if 'переработка:' in i:
            #print(i)
            pass
        else:
            #print(i,file=sys.stderr)
            pass
    else:
        pass
print('---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------')
for n in list_print_2:
    #print(n)
    pass
#-------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
list_wage = []
info_wage = {}


for dt in daterange(a, b, inclusive=True):
    date = dt.strftime("%Y-%m-%d")
    if date in full_info.keys():
        for i in full_info[date].keys():
                if not i in info_wage.keys():
                        info_wage[i] = {}
                elif i in info_wage.keys():
                        if not date in info_wage[i].keys():
                                info_wage[i][date]=[full_info[date][i][2],full_info[date][i][3]]
                else:   
                        pass
recycling_all_money = 0
recycling_all_time = 0
for i in number_family:
        if i in info_wage.keys():
                days = len(info_wage[i])
                for dt in daterange(a, b, inclusive=True):
                        date = dt.strftime("%Y-%m-%d")
                        for date in info_wage[i].keys():
                                recycling_money = info_wage[i][date][1]*money_rate_1s
                                recycling_all_money = recycling_all_money
                                recycling_all_time = recycling_all_time + info_wage[i][date][1]
                recycling_all_time_str = time.strftime('%H:%M:%S', time.gmtime(recycling_all_time))
                wage_all = days*money_rate_8h + recycling_all_money
                wage_all_round = round(wage_all,3)
                name = replace_name(i)
                print('\n\n\t',name,'\nЗарплата:  ',wage_all,'\n\nОбщаяя переработка: ', recycling_all_time_str)
                print('\n\n\n\n\n',recycling_all_time)


#-------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
print('FINAL!')
#print(full_info)
esc_out = input()
