from secondary_func import*
from datetime import timedelta
from datetime import datetime
import time
import sys
import openpyxl
wb = openpyxl.Workbook()
ws = wb.active
topicsList={'Фамилия', 'Дата', 'Отметка входа', 'Отметка выхода', 'Общее время работы', 'Переботка', 'Примечания'}
topicCounter=1
for topic in topicsList:
    ws.cell(column = topicCounter, row = 1, value = topic)
    topicCounter+=1
full_info = {}
time_str = '00:00:00'
count = 2 #Начальное значение счетчика. С помощью этой переменной определяется значеение счетчика 
delta_work = datetime.strptime(time_str,"%H:%M:%S")# Реально отработанное время
work_time = timedelta(hours=8)# Длительность рабочего дня
time_comparison = timedelta(seconds=0) #Разница между реально отработанным временем и длительностью рабочего дня
recycling_seconds = 0
recycling_hour = 0.0
recycling_hour_round = 0.0
tag = '0'
recycling_data = time.strftime('%H:%M:%S', time.gmtime(abs(recycling_seconds)))
f = open('resources/12_2018','r')# Иеальные данные tst_0/Не идеальные данные tst_1 Загрузка файла с данными
for line in f:# Перебор всех строк в файле с данными
    Y = line[8:29]
    date_and_time = datetime.strptime(line[10:29], "%Y-%m-%d %H:%M:%S")
    date = line[10:20]
    U = int(Y[0])
    family = replace_name(U)
    if not  date in full_info.keys():
        full_info[date] = {}
    if U in full_info[date].keys():
        if full_info[date][U][0] < date_and_time:
            full_info[date][U][0] = date_and_time
        if full_info[date][U][1] > date_and_time:
            full_info[date][U][1] = date_and_time
    else:
        full_info[date][U] = [date_and_time, date_and_time,delta_work,recycling_seconds,recycling_hour,recycling_hour_round,recycling_data,tag]
number_family = [1,2,3,4,5,6,7,8,9,10] ; #id работников 
year = input("Get a year: ") # ввод года для построения табеля
month = input("Get a month: ") # ввод месяц для построения табеля
last_day = monthrange(int(year), int(month))[1] # Определение последнего дня месяца? 
a = datetime(int(year), int(month), 1) ;
b = datetime(int(year), int(month), last_day)
for dt in daterange(a, b, inclusive=True):
  date = dt.strftime("%Y-%m-%d")
  if date in full_info.keys():
    for i in full_info[date].keys():
      name = replace_name(i)
      rec_name = ws.cell(column = 1, row = count, value = name)
      rec_date = ws.cell(column = 2, row = count, value = date)
      if name != 'Рощин':
        if full_info[date][i][0] == full_info[date][i][1]:
          print("ПРЕДУПРЕЖДЕНИЕ! " + name + " имеет только одну отметку в рабочем дне!", file=sys.stderr)
          print('Дата время в системе',full_info[date][i][0])
          print('\n\n\tВыберете вариант:\n\n\t1.Ввести время прихода\n\t2.Введите время ухода')
          l = 0
          while l!=1:
            time_inp = input('\n\n\tВведите пункт меню:')
            if time_inp == '1' or time_inp == '2':
              l = 1
              if time_inp == '1':
                time_begin = input('Введите время прихода в формате час*ПРОБЕЛ*минуты*секунды(если есть)')
                time_begin_str = date + ' ' + time_begin
                time_begin_z = datetime.strptime(time_begin_str,"%Y-%m-%d %H %M %S")
                full_info[date][i][0] = time_begin_z
              elif time_inp == '2': 
                time_end = input('Введите время ухода в формате час*ПРОБЕЛ*минуты*секунды(если есть)')
                time_end_str = date + ' ' + time_end
                time_end_z = datetime.strptime(time_end_str,"%Y-%m-%d %H %M %S")
                full_info[date][i][1] = time_end_z
              else:
                pass
              print('после измений')
              value_for_record = datetime.time(full_info[date][i][0])
              rec_time_start = ws.cell(column = 3, row = count, value = value_for_record)
              value_for_record = datetime.time(full_info[date][i][1])
              rec_time_end = ws.cell(column = 4, row = count, value = value_for_record)
              work_delta = full_info[date][i][1] - full_info[date][i][0]
              work_delta_seconds = (full_info[date][i][1] - full_info[date][i][0]).seconds
              full_info[date][i][2] = work_delta
              rec_delta = ws.cell(column = 5, row = count, value = full_info[date][i][2])
              recycling_seconds = work_delta_seconds - work_time.seconds
              recycling_hour = recycling_seconds/3600
              recycling_hour_round = round(recycling_hour,2)
              recycling_data = time.strftime('%H:%M:%S', time.gmtime(abs(recycling_seconds)))
              full_info[date][i][3] = recycling_seconds
              full_info[date][i][4] = recycling_hour
              full_info[date][i][5] = recycling_hour_round
              full_info[date][i][6] = recycling_data
              auto_tag = 'auto'
              full_info[date][i][7] = auto_tag
              if full_info[date][i][3] >= 0:
                info_auto = 'Ручной ввод'
                rec_recycling = ws.cell(column = 7, row = count, value = info_auto)
                rec_recycling = ws.cell(column = 6, row = count, value = full_info[date][i][6])
              else:
                info_auto = 'Ручной ввод'
                rec_recycling = ws.cell(column = 7, row = count, value = info_auto)
                info_1 = '- ' + str(full_info[date][i][6])
                rec_recycling = ws.cell(column = 6, row = count, value = info_1)
              auto_tag = 'no_auto'
              full_info[date][i][7] = auto_tag
          else:
            print('Введите правильный пункт меню!')
        else:
          work_delta = full_info[date][i][0] - full_info[date][i][1]#Короче весь этот блок надо переделывать. Полность. Считать через секунды и формировать значения из  секунд. Еще надо сменить название переменной переботки на более корректное. Добавить id новых сотрудинов, выкинуть Федоровича из программы. Разобраться с форматированнием вывода времени и даты. Разобраться с ранжированием месяца.
          work_delta_seconds = work_delta.seconds
          value_for_record = datetime.time(full_info[date][i][0])
          rec_time_start = ws.cell(column = 4, row = count, value = value_for_record)
          value_for_record = datetime.time(full_info[date][i][1])
          rec_time_end = ws.cell(column = 3, row = count, value = value_for_record)
          full_info[date][i][2] = work_delta
          rec_delta = ws.cell(column = 5, row = count, value = full_info[date][i][2])
          recycling_seconds = work_delta_seconds - work_time.seconds
          recycling_hour = recycling_seconds/3600
          recycling_hour_round = round(recycling_hour,2) 
          recycling_data = time.strftime('%H:%M:%S', time.gmtime(abs(recycling_seconds)))
          full_info[date][i][3] = recycling_seconds
          full_info[date][i][4] = recycling_hour
          full_info[date][i][5] = recycling_hour_round
          full_info[date][i][6] = recycling_data
          auto_tag = 'auto'
          full_info[date][i][7] = auto_tag
          if full_info[date][i][3] >= 0:
                info_auto = 'Автоматика'
                rec_recycling = ws.cell(column = 7, row = count, value = info_auto)
                rec_recycling = ws.cell(column = 6, row = count, value = full_info[date][i][6])
          else:
                info_auto = 'Автоматика'
                rec_recycling = ws.cell(column = 7, row = count, value = info_auto)
                info_1 = '- ' + str(full_info[date][i][6])
                rec_recycling = ws.cell(column = 6, row = count, value = info_1)
          count = count +1
      else:
          value_for_record = datetime.time(full_info[date][i][0])
          rec_time_start = ws.cell(column = 3, row = count, value = value_for_record)
          auto_tag = 0
          full_info[date][i][7] = auto_tag
          count = count +1
    else:
        pass
    

wb.save("Декабрь_2018_вариант_1.xlsx")
#-------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
list_print_1 = []
list_print_2 = []
for dt in daterange(a, b, inclusive=True):
    date = dt.strftime("%Y-%m-%d")
    if date in full_info.keys():
        for i in full_info[date].keys():
            name = replace_name(i)
            if name != 'Рощин':
                if full_info[date][i][3] >= 0:
                    prt_1 = str(name)+ ' ' + str(date) + ' '+ str(full_info[date][i][2])+ ' ' + ' переработка: ' + str(full_info[date][i][6])
                else:
                    prt_1 = str(name)+ ' ' + str(date) + ' '+ str(full_info[date][i][2])+ ' ' + ' недоработка: ' + str(full_info[date][i][6])
                list_print_1.append(prt_1)
            else:
                prt_2 = str(name)+ ' ' + str(date) + ' '+ 'время прихода' + str(full_info[date][i][0])
                list_print_2.append(prt_2)
for i in list_print_1:
    if not 'Еремин' in i:
        if 'переработка:' in i:
            print(i)
        else:
            print(i,file=sys.stderr)
    else:
        pass
print('---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------')
for n in list_print_2:
    print(n)
print('FINAL!')
#-------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
count = 2
wb_1 = openpyxl.Workbook()
ws_1 = wb_1.active
ws_1.cell(column = 1, row = 1, value = 'Фамилия')
ws_1.cell(column = 2, row = 1, value = 'Дата')
ws_1.cell(column = 3, row = 1, value = 'Отметка входа')
ws_1.cell(column = 4, row = 1, value = 'Отметка выхода')
ws_1.cell(column = 5, row = 1, value = 'Общее время работы')
ws_1.cell(column = 6, row = 1, value = 'Переботка')
ws_1.cell(column = 7, row = 1, value = 'Примечания')
for dt in daterange(a, b, inclusive=True):
    date = dt.strftime("%Y-%m-%d")
    if date in full_info.keys():
        for i in full_info[date].keys():
            name = replace_name(i)
            rec_name = ws_1.cell(column = 1, row = count, value = name)
            rec_date = ws_1.cell(column = 2, row = count, value = date)
            if name != 'Рощин':
                value_for_record = datetime.time(full_info[date][i][0])
                rec_time_start = ws_1.cell(column = 3, row = count, value = value_for_record)
                value_for_record = datetime.time(full_info[date][i][1])
                rec_time_end = ws_1.cell(column = 4, row = count, value = value_for_record)
                rec_delta = ws_1.cell(column = 5, row = count, value = full_info[date][i][2])
                if full_info[date][i][3] >= 0:
                        if full_info[date][i][7]!= 'auto':
                                info_auto = 'Ручной ввод'
                                rec_recycling = ws_1.cell(column = 7, row = count, value = info_auto)
                                print(full_info[date][i][6],'Переработка')
                                rec_recycling = ws_1.cell(column = 6, row = count, value = full_info[date][i][6])
                        else:
                             info_auto = 'Автоматика'
                             rec_recycling = ws_1.cell(column = 7, row = count, value = info_auto)
                             print(full_info[date][i][6],'Переработка')
                             rec_recycling = ws_1.cell(column = 6, row = count, value = full_info[date][i][6])
                else:
                        if full_info[date][i][7]!= 'auto':
                                info_auto = 'Ручной ввод'
                                info_1 = '- ' + str(full_info[date][i][6])
                                rec_recycling = ws_1.cell(column = 7, row = count, value = info_auto)
                                print(full_info[date][i][6],'Переработка')
                                rec_recycling = ws_1.cell(column = 6, row = count, value = info_1)
                        else:
                             info_auto = 'Автоматика'
                             info_1 = '- ' + str(full_info[date][i][6])
                             rec_recycling = ws_1.cell(column = 7, row = count, value = info_auto)
                             print(full_info[date][i][6],'Переработка')
                             rec_recycling = ws_1.cell(column = 6, row = count, value = info_1)
            else:
                value_for_record = datetime.time(full_info[date][i][0])
                rec_time_start = ws_1.cell(column = 3, row = count, value = value_for_record)
            count += 1
print('----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------')
wb_1.save("Декабрь_2018_вариант_2.xlsx")
print('Файл сформирован')
for n in list_print_2:
    print(n)
print('FINAL!')
esc_out = input()
